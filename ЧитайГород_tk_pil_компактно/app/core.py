from __future__ import annotations

import hashlib
import re
import shutil
import sqlite3
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from PIL import Image, ImageOps

BASE_DIR = Path(__file__).resolve().parents[1]
IMPORT_DIR = BASE_DIR / "import"
ASSETS_DIR = BASE_DIR / "assets"
PRODUCT_IMAGES_DIR = ASSETS_DIR / "products"
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "bookstore.db"
PLACEHOLDER_REL = "picture.png"
PLACEHOLDER_PATH = ASSETS_DIR / PLACEHOLDER_REL
ICON_PATH = ASSETS_DIR / "icon.png"
ICON_ICO_PATH = ASSETS_DIR / "icon.ico"

ROLE_ADMIN = "Администратор"
ROLE_MANAGER = "Менеджер"
ROLE_CLIENT = "Авторизированный клиент"
ROLES = {ROLE_ADMIN, ROLE_MANAGER, ROLE_CLIENT}

STATUS_OPTIONS = ["Новый", "В работе", "Завершен", "Отменен"]
DISCOUNT_FILTERS = ["Все диапазоны", "0-12,99%", "13-16,99%", "17% и более"]

ALLOWED_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}

SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    role TEXT NOT NULL,
    full_name TEXT NOT NULL,
    login TEXT NOT NULL UNIQUE,
    password_hash TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS categories (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS suppliers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS manufacturers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS pickup_points (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    address TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    article TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    unit TEXT NOT NULL,
    price NUMERIC NOT NULL,
    discount INTEGER NOT NULL DEFAULT 0,
    stock_qty INTEGER NOT NULL DEFAULT 0,
    description TEXT NOT NULL DEFAULT '',
    photo_path TEXT NOT NULL,
    category_id INTEGER NOT NULL,
    supplier_id INTEGER NOT NULL,
    manufacturer_id INTEGER NOT NULL,
    FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE RESTRICT,
    FOREIGN KEY(supplier_id) REFERENCES suppliers(id) ON DELETE RESTRICT,
    FOREIGN KEY(manufacturer_id) REFERENCES manufacturers(id) ON DELETE RESTRICT
);

CREATE TABLE IF NOT EXISTS orders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    order_number INTEGER NOT NULL UNIQUE,
    status TEXT NOT NULL,
    order_date TEXT,
    delivery_date TEXT,
    receive_code INTEGER NOT NULL,
    customer_id INTEGER NOT NULL,
    pickup_point_id INTEGER NOT NULL,
    FOREIGN KEY(customer_id) REFERENCES users(id) ON DELETE RESTRICT,
    FOREIGN KEY(pickup_point_id) REFERENCES pickup_points(id) ON DELETE RESTRICT
);

CREATE TABLE IF NOT EXISTS order_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id INTEGER NOT NULL,
    product_id INTEGER NOT NULL,
    quantity INTEGER NOT NULL,
    UNIQUE(order_id, product_id),
    FOREIGN KEY(order_id) REFERENCES orders(id) ON DELETE CASCADE,
    FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE RESTRICT
);
"""


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def slugify(value: str) -> str:
    value = normalize_text(value).lower()
    value = re.sub(r"[^a-z0-9а-яё]+", "_", value, flags=re.IGNORECASE)
    value = re.sub(r"_+", "_", value).strip("_")
    return value or "file"


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def verify_password(password: str, stored_hash: str) -> bool:
    return hash_password(password) == stored_hash


def format_money(value: object) -> str:
    try:
        dec = Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError):
        dec = Decimal("0.00")
    return f"{dec:.2f}"


def parse_decimal(value: str) -> Decimal:
    text = normalize_text(value).replace(",", ".")
    if not text:
        raise ValueError("Поле «Цена» обязательно для заполнения.")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("Введите корректную цену.") from exc
    if number < 0:
        raise ValueError("Цена не может быть отрицательной.")
    return number.quantize(Decimal("0.01"))


def parse_int(value: str, field_name: str, min_value: int = 0, max_value: int | None = None) -> int:
    text = normalize_text(value)
    if not text:
        raise ValueError(f"Поле «{field_name}» обязательно для заполнения.")
    if not re.fullmatch(r"-?\d+", text):
        raise ValueError(f"Поле «{field_name}» должно быть целым числом.")
    number = int(text)
    if number < min_value:
        raise ValueError(f"Поле «{field_name}» не может быть меньше {min_value}.")
    if max_value is not None and number > max_value:
        raise ValueError(f"Поле «{field_name}» не может быть больше {max_value}.")
    return number


def parse_iso_date(value: str, field_name: str, allow_empty: bool = True) -> str | None:
    text = normalize_text(value)
    if not text:
        if allow_empty:
            return None
        raise ValueError(f"Поле «{field_name}» обязательно для заполнения.")
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        raise ValueError(f"Поле «{field_name}» должно быть в формате ГГГГ-ММ-ДД.")
    return text


@dataclass(slots=True)
class ProductView:
    id: int
    article: str
    name: str
    unit: str
    price: Decimal
    final_price: Decimal
    discount: int
    stock_qty: int
    description: str
    category: str
    supplier: str
    manufacturer: str
    photo_path: str

    @property
    def search_text(self) -> str:
        return " ".join(
            [
                self.article,
                self.name,
                self.unit,
                self.description,
                self.category,
                self.supplier,
                self.manufacturer,
                format_money(self.price),
                str(self.discount),
                str(self.stock_qty),
            ]
        ).lower()

    @property
    def is_discount_highlight(self) -> bool:
        return self.discount > 25

    @property
    def is_out_of_stock(self) -> bool:
        return self.stock_qty <= 0


@dataclass(slots=True)
class OrderItemView:
    product_id: int
    article: str
    quantity: int


@dataclass(slots=True)
class OrderView:
    id: int
    order_number: int
    status: str
    order_date: str | None
    delivery_date: str | None
    receive_code: int
    customer_name: str
    pickup_point: str
    items: list[OrderItemView]

    @property
    def articles_text(self) -> str:
        return "; ".join(f"{item.article} × {item.quantity}" for item in self.items)

    @property
    def search_text(self) -> str:
        base = [
            str(self.order_number),
            self.status,
            self.customer_name,
            self.pickup_point,
            self.order_date or "",
            self.delivery_date or "",
            self.articles_text,
            str(self.receive_code),
        ]
        return " ".join(base).lower()


@dataclass(slots=True)
class SessionUser:
    id: int | None
    role: str
    full_name: str

    @property
    def is_guest(self) -> bool:
        return self.id is None and self.role == "Гость"



def connect(db_path: Path | None = None) -> sqlite3.Connection:
    path = db_path or DB_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn



def ensure_assets() -> None:
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    PRODUCT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    for file_name in ("icon.png", "icon.ico", "picture.png"):
        src = IMPORT_DIR / file_name
        dst = ASSETS_DIR / file_name
        if src.exists() and not dst.exists():
            shutil.copy2(src, dst)



def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(SCHEMA_SQL)
    conn.commit()



def _read_sheet_rows(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [normalize_text(h) for h in rows[0]]
    result: list[dict[str, object]] = []
    for row in rows[1:]:
        if not row or all(cell in (None, "") for cell in row):
            continue
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        if any(normalize_text(v) for v in item.values()):
            result.append(item)
    return result



def _read_single_column_rows(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    result: list[str] = []
    for row in ws.iter_rows(values_only=True):
        value = normalize_text(row[0] if row else "")
        if value:
            result.append(value)
    return result



def _copy_and_resize_image(source: Path, dest_name: str) -> str:
    ext = source.suffix.lower()
    if ext not in ALLOWED_IMAGE_EXTS:
        ext = ".jpg"
    dest_path = PRODUCT_IMAGES_DIR / dest_name
    with Image.open(source) as img:
        image = ImageOps.fit(img.convert("RGB"), (300, 200), method=Image.Resampling.LANCZOS)
        image.save(dest_path, quality=90)
    return f"products/{dest_path.name}"



def _placeholder_rel() -> str:
    return PLACEHOLDER_REL



def _ensure_data_imported(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] > 0:
        return

    ensure_assets()

    users_rows = _read_sheet_rows(IMPORT_DIR / "user_import.xlsx")
    product_rows = _read_sheet_rows(IMPORT_DIR / "Tovar.xlsx")
    order_rows = _read_sheet_rows(IMPORT_DIR / "Заказ_import.xlsx")
    pickup_points = _read_single_column_rows(IMPORT_DIR / "Пункты выдачи_import.xlsx")

    category_ids: dict[str, int] = {}
    supplier_ids: dict[str, int] = {}
    manufacturer_ids: dict[str, int] = {}
    customer_ids_by_name: dict[str, int] = {}

    for row in users_rows:
        role = normalize_text(row.get("Роль сотрудника"))
        full_name = normalize_text(row.get("ФИО"))
        login = normalize_text(row.get("Логин"))
        password = normalize_text(row.get("Пароль"))
        conn.execute(
            "INSERT INTO users(role, full_name, login, password_hash) VALUES (?, ?, ?, ?)",
            (role, full_name, login, hash_password(password)),
        )
        customer_ids_by_name[full_name] = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()

    for name in sorted({normalize_text(r.get("Категория товара")) for r in product_rows}):
        cur = conn.execute("INSERT OR IGNORE INTO categories(name) VALUES (?)", (name,))
        if cur.rowcount == 1:
            category_ids[name] = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        else:
            category_ids[name] = int(conn.execute("SELECT id FROM categories WHERE name=?", (name,)).fetchone()[0])

    for name in sorted({normalize_text(r.get("Поставщик")) for r in product_rows}):
        cur = conn.execute("INSERT OR IGNORE INTO suppliers(name) VALUES (?)", (name,))
        if cur.rowcount == 1:
            supplier_ids[name] = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        else:
            supplier_ids[name] = int(conn.execute("SELECT id FROM suppliers WHERE name=?", (name,)).fetchone()[0])

    for name in sorted({normalize_text(r.get("Производитель")) for r in product_rows}):
        cur = conn.execute("INSERT OR IGNORE INTO manufacturers(name) VALUES (?)", (name,))
        if cur.rowcount == 1:
            manufacturer_ids[name] = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        else:
            manufacturer_ids[name] = int(conn.execute("SELECT id FROM manufacturers WHERE name=?", (name,)).fetchone()[0])

    conn.commit()

    for row in product_rows:
        photo_name = normalize_text(row.get("Фото"))
        image_rel = _placeholder_rel()
        source_image = IMPORT_DIR / photo_name if photo_name else None
        if source_image and source_image.exists():
            image_rel = f"import/{photo_name}"
        conn.execute(
            """
            INSERT INTO products(article, name, unit, price, discount, stock_qty, description, photo_path, category_id, supplier_id, manufacturer_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                normalize_text(row.get("Артикул")),
                normalize_text(row.get("Наименование товара")),
                normalize_text(row.get("Единица измерения")),
                float(Decimal(str(row.get("Цена"))).quantize(Decimal("0.01"))),
                int(row.get("Действующая скидка") or 0),
                int(row.get("Кол-во на складе") or 0),
                normalize_text(row.get("Описание товара")),
                image_rel,
                category_ids[normalize_text(row.get("Категория товара"))],
                supplier_ids[normalize_text(row.get("Поставщик"))],
                manufacturer_ids[normalize_text(row.get("Производитель"))],
            ),
        )
    conn.commit()

    for address in pickup_points:
        conn.execute("INSERT INTO pickup_points(address) VALUES (?)", (address,))
    conn.commit()

    product_ids_by_article = {row["article"]: row["id"] for row in conn.execute("SELECT id, article FROM products").fetchall()}
    pickup_ids = [row["id"] for row in conn.execute("SELECT id FROM pickup_points ORDER BY id").fetchall()]

    next_code = 900
    for row in order_rows:
        order_number = int(row.get("Номер заказа") or 0)
        customer_name = normalize_text(row.get("ФИО авторизированного клиента"))
        customer_id = customer_ids_by_name.get(customer_name)
        if customer_id is None:
            continue
        pickup_index = int(row.get("Адрес пункта выдачи") or 1)
        pickup_point_id = pickup_ids[pickup_index - 1] if 1 <= pickup_index <= len(pickup_ids) else pickup_ids[0]
        order_date = _to_iso_date(row.get("Дата заказа"))
        delivery_date = _to_iso_date(row.get("Дата доставки"))
        status = normalize_text(row.get("Статус заказа")) or "Новый"
        receive_code = int(row.get("Код для получения") or next_code)
        next_code += 1
        cur = conn.execute(
            """
            INSERT INTO orders(order_number, status, order_date, delivery_date, receive_code, customer_id, pickup_point_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (order_number, status, order_date, delivery_date, receive_code, customer_id, pickup_point_id),
        )
        order_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        raw = normalize_text(row.get("Артикул заказа"))
        parts = [p.strip() for p in re.split(r"[,;\n]", raw) if p.strip()]
        i = 0
        while i < len(parts):
            article = parts[i]
            quantity = 1
            if i + 1 < len(parts) and re.fullmatch(r"\d+", parts[i + 1]):
                quantity = int(parts[i + 1])
                i += 2
            else:
                i += 1
            product_id = product_ids_by_article.get(article)
            if product_id:
                conn.execute(
                    "INSERT OR IGNORE INTO order_items(order_id, product_id, quantity) VALUES (?, ?, ?)",
                    (order_id, product_id, quantity),
                )
        conn.commit()



def _to_iso_date(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    try:
        from datetime import datetime

        parsed = datetime.fromisoformat(text)
        return parsed.date().isoformat()
    except Exception:
        try:
            from pandas import to_datetime  # type: ignore

            parsed = to_datetime(value, errors="coerce", dayfirst=True)
            if parsed is not None and not getattr(parsed, "isna", lambda: True)():
                return parsed.date().isoformat()
        except Exception:
            return None
    return None



def initialize_database(db_path: Path | None = None) -> Path:
    path = db_path or DB_PATH
    conn = connect(path)
    create_schema(conn)
    _ensure_data_imported(conn)
    conn.close()
    return path



def _product_row_to_view(row: sqlite3.Row) -> ProductView:
    price = Decimal(str(row["price"])).quantize(Decimal("0.01"))
    final_price = (price * (Decimal("1") - Decimal(row["discount"]) / Decimal("100"))).quantize(Decimal("0.01")) if int(row["discount"]) > 0 else price
    return ProductView(
        id=row["id"],
        article=row["article"],
        name=row["name"],
        unit=row["unit"],
        price=price,
        final_price=final_price,
        discount=int(row["discount"]),
        stock_qty=int(row["stock_qty"]),
        description=row["description"],
        category=row["category_name"],
        supplier=row["supplier_name"],
        manufacturer=row["manufacturer_name"],
        photo_path=row["photo_path"],
    )



def get_products(
    conn: sqlite3.Connection,
    search: str = "",
    sort_key: str = "id",
    sort_dir: str = "asc",
    discount_filter: str = "Все диапазоны",
) -> list[ProductView]:
    rows = conn.execute(
        """
        SELECT p.*, c.name AS category_name, s.name AS supplier_name, m.name AS manufacturer_name
        FROM products p
        JOIN categories c ON c.id = p.category_id
        JOIN suppliers s ON s.id = p.supplier_id
        JOIN manufacturers m ON m.id = p.manufacturer_id
        ORDER BY p.id ASC
        """
    ).fetchall()
    products = [_product_row_to_view(row) for row in rows]

    query = normalize_text(search).lower()
    if query:
        tokens = [t for t in re.split(r"\s+", query) if t]
        filtered: list[ProductView] = []
        for product in products:
            text = product.search_text
            if all(token in text for token in tokens):
                filtered.append(product)
        products = filtered

    if discount_filter == "0-12,99%":
        products = [p for p in products if 0 <= p.discount <= 12]
    elif discount_filter == "13-16,99%":
        products = [p for p in products if 13 <= p.discount <= 16]
    elif discount_filter == "17% и более":
        products = [p for p in products if p.discount >= 17]

    reverse = sort_dir == "desc"
    key_map = {
        "price": lambda p: (p.price, p.id),
        "stock_qty": lambda p: (p.stock_qty, p.id),
        "discount": lambda p: (p.discount, p.id),
        "name": lambda p: (p.name.lower(), p.id),
        "id": lambda p: (p.id,),
    }
    products.sort(key=key_map.get(sort_key, key_map["id"]), reverse=reverse)
    return products



def get_orders(conn: sqlite3.Connection, search: str = "") -> list[OrderView]:
    order_rows = conn.execute(
        """
        SELECT o.*, u.full_name AS customer_name, pp.address AS pickup_point
        FROM orders o
        JOIN users u ON u.id = o.customer_id
        JOIN pickup_points pp ON pp.id = o.pickup_point_id
        ORDER BY o.order_number ASC
        """
    ).fetchall()
    result: list[OrderView] = []
    for row in order_rows:
        items_rows = conn.execute(
            """
            SELECT oi.product_id, oi.quantity, p.article
            FROM order_items oi
            JOIN products p ON p.id = oi.product_id
            WHERE oi.order_id = ?
            ORDER BY oi.id ASC
            """,
            (row["id"],),
        ).fetchall()
        items = [OrderItemView(product_id=r["product_id"], article=r["article"], quantity=int(r["quantity"])) for r in items_rows]
        result.append(
            OrderView(
                id=row["id"],
                order_number=int(row["order_number"]),
                status=row["status"],
                order_date=row["order_date"],
                delivery_date=row["delivery_date"],
                receive_code=int(row["receive_code"]),
                customer_name=row["customer_name"],
                pickup_point=row["pickup_point"],
                items=items,
            )
        )
    query = normalize_text(search).lower()
    if query:
        tokens = [t for t in re.split(r"\s+", query) if t]
        result = [order for order in result if all(token in order.search_text for token in tokens)]
    return result



def get_reference_data(conn: sqlite3.Connection) -> dict[str, list[str]]:
    data = {
        "categories": [row[0] for row in conn.execute("SELECT name FROM categories ORDER BY name").fetchall()],
        "suppliers": [row[0] for row in conn.execute("SELECT name FROM suppliers ORDER BY name").fetchall()],
        "manufacturers": [row[0] for row in conn.execute("SELECT name FROM manufacturers ORDER BY name").fetchall()],
        "pickup_points": [row[0] for row in conn.execute("SELECT address FROM pickup_points ORDER BY id").fetchall()],
        "customers": [row[0] for row in conn.execute("SELECT full_name FROM users WHERE role = ? ORDER BY full_name", (ROLE_CLIENT,)).fetchall()],
    }
    return data



def next_product_article_hint(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM products").fetchone()
    return int(row[0])



def next_order_number_hint(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT COALESCE(MAX(order_number), 0) + 1 FROM orders").fetchone()
    return int(row[0])



def _make_product_photo_path(product_id: int, source_file: Path) -> str:
    ext = source_file.suffix.lower() if source_file.suffix.lower() in ALLOWED_IMAGE_EXTS else ".jpg"
    filename = f"product_{product_id}{ext}"
    return f"products/{filename}"


def _resolve_product_image_path(photo_path: str) -> Path:
    rel = normalize_text(photo_path) or PLACEHOLDER_REL
    rel_path = Path(rel)
    candidates: list[Path] = []
    if rel_path.parts and rel_path.parts[0] == "products":
        candidates.append(ASSETS_DIR / rel_path)
        candidates.append(IMPORT_DIR / rel_path.name)
    else:
        candidates.append(ASSETS_DIR / rel_path)
        candidates.append(IMPORT_DIR / rel_path.name)
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return PLACEHOLDER_PATH



def save_product(
    conn: sqlite3.Connection,
    *,
    product_id: int | None,
    article: str,
    name: str,
    category: str,
    supplier: str,
    manufacturer: str,
    unit: str,
    price: str,
    stock_qty: str,
    discount: str,
    description: str,
    photo_source: str | None,
) -> int:
    article = normalize_text(article)
    name = normalize_text(name)
    category = normalize_text(category)
    supplier = normalize_text(supplier)
    manufacturer = normalize_text(manufacturer)
    unit = normalize_text(unit)
    description = normalize_text(description)
    if not article:
        raise ValueError("Поле «Артикул» обязательно для заполнения.")
    if not name:
        raise ValueError("Поле «Наименование» обязательно для заполнения.")
    if not category:
        raise ValueError("Выберите категорию товара.")
    if not supplier:
        raise ValueError("Выберите поставщика.")
    if not manufacturer:
        raise ValueError("Выберите производителя.")
    if not unit:
        raise ValueError("Поле «Единица измерения» обязательно для заполнения.")

    price_value = parse_decimal(price)
    stock_value = parse_int(stock_qty, "Количество на складе", 0)
    discount_value = parse_int(discount, "Действующая скидка", 0, 100)
    if discount_value > 100:
        raise ValueError("Скидка не может превышать 100%.")

    category_id = conn.execute("SELECT id FROM categories WHERE name = ?", (category,)).fetchone()
    supplier_id = conn.execute("SELECT id FROM suppliers WHERE name = ?", (supplier,)).fetchone()
    manufacturer_id = conn.execute("SELECT id FROM manufacturers WHERE name = ?", (manufacturer,)).fetchone()
    if not all([category_id, supplier_id, manufacturer_id]):
        raise ValueError("Не удалось найти связанный справочник. Обновите список и повторите попытку.")

    photo_path = PLACEHOLDER_REL
    old_photo_path = None
    if product_id is None:
        cur = conn.execute(
            """
            INSERT INTO products(article, name, unit, price, discount, stock_qty, description, photo_path, category_id, supplier_id, manufacturer_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (article, name, unit, float(price_value), discount_value, stock_value, description, photo_path, category_id[0], supplier_id[0], manufacturer_id[0]),
        )
        product_id = int(cur.lastrowid)
    else:
        row = conn.execute("SELECT photo_path FROM products WHERE id = ?", (product_id,)).fetchone()
        if row is None:
            raise ValueError("Товар не найден.")
        old_photo_path = row[0]
        conn.execute(
            """
            UPDATE products
            SET article = ?, name = ?, unit = ?, price = ?, discount = ?, stock_qty = ?, description = ?, category_id = ?, supplier_id = ?, manufacturer_id = ?
            WHERE id = ?
            """,
            (article, name, unit, float(price_value), discount_value, stock_value, description, category_id[0], supplier_id[0], manufacturer_id[0], product_id),
        )

    if photo_source:
        source_path = Path(photo_source)
        if not source_path.exists():
            raise ValueError("Выбранный файл изображения не найден.")
        new_photo_path = _make_product_photo_path(product_id, source_path)
        final_path = ASSETS_DIR / new_photo_path
        with Image.open(source_path) as img:
            image = ImageOps.fit(img.convert("RGB"), (300, 200), method=Image.Resampling.LANCZOS)
            image.save(final_path, quality=90)
        conn.execute("UPDATE products SET photo_path = ? WHERE id = ?", (new_photo_path, product_id))
        if old_photo_path and old_photo_path != PLACEHOLDER_REL and old_photo_path != new_photo_path:
            old_path = ASSETS_DIR / old_photo_path
            if old_path.exists() and old_path.is_file():
                try:
                    old_path.unlink()
                except OSError:
                    pass
    conn.commit()
    return product_id



def delete_product(conn: sqlite3.Connection, product_id: int) -> None:
    in_use = conn.execute("SELECT 1 FROM order_items WHERE product_id = ? LIMIT 1", (product_id,)).fetchone()
    if in_use:
        raise ValueError("Удаление невозможно: товар используется в одном или нескольких заказах.")
    row = conn.execute("SELECT photo_path FROM products WHERE id = ?", (product_id,)).fetchone()
    if row is None:
        raise ValueError("Товар не найден.")
    photo_path = row[0]
    conn.execute("DELETE FROM products WHERE id = ?", (product_id,))
    conn.commit()
    if photo_path and photo_path != PLACEHOLDER_REL:
        path = ASSETS_DIR / photo_path
        if path.exists() and path.is_file():
            try:
                path.unlink()
            except OSError:
                pass



def _parse_order_articles(raw_articles: str) -> list[tuple[str, int]]:
    text = normalize_text(raw_articles)
    if not text:
        return []
    normalized = text.replace("\n", ";")
    chunks = [chunk.strip() for chunk in re.split(r"[;]+", normalized) if chunk.strip()]
    result: list[tuple[str, int]] = []
    for chunk in chunks:
        m = re.fullmatch(r"(.+?)\s*[×xX]\s*(\d+)", chunk)
        if m:
            article = normalize_text(m.group(1))
            qty = int(m.group(2))
            result.append((article, qty))
            continue
        parts = [p.strip() for p in re.split(r"[,]+", chunk) if p.strip()]
        if len(parts) >= 2 and re.fullmatch(r"\d+", parts[1]):
            result.append((parts[0], int(parts[1])))
        else:
            result.append((chunk, 1))
    return result



def save_order(
    conn: sqlite3.Connection,
    *,
    order_id: int | None,
    order_number: str,
    status: str,
    pickup_point: str,
    order_date: str,
    delivery_date: str,
    customer: str,
    articles_text: str,
    receive_code: str | None = None,
) -> int:
    order_number_value = parse_int(order_number, "Номер заказа", 1)
    status = normalize_text(status)
    pickup_point = normalize_text(pickup_point)
    customer = normalize_text(customer)
    articles_text = normalize_text(articles_text)
    order_date_value = parse_iso_date(order_date, "Дата заказа", allow_empty=True)
    delivery_date_value = parse_iso_date(delivery_date, "Дата выдачи", allow_empty=True)
    if not status:
        raise ValueError("Выберите статус заказа.")
    if not pickup_point:
        raise ValueError("Выберите адрес пункта выдачи.")
    if not customer:
        raise ValueError("Выберите клиента.")
    if not articles_text:
        raise ValueError("Укажите артикулы товара.")

    pickup_row = conn.execute("SELECT id FROM pickup_points WHERE address = ?", (pickup_point,)).fetchone()
    customer_row = conn.execute("SELECT id FROM users WHERE full_name = ?", (customer,)).fetchone()
    if pickup_row is None:
        raise ValueError("Не найден пункт выдачи.")
    if customer_row is None:
        raise ValueError("Не найден клиент.")

    articles = _parse_order_articles(articles_text)
    if not articles:
        raise ValueError("Укажите хотя бы один артикул товара.")

    items: list[tuple[int, int]] = []
    for article, qty in articles:
        product_row = conn.execute("SELECT id FROM products WHERE article = ?", (article,)).fetchone()
        if product_row is None:
            raise ValueError(f"Товар с артикулом {article} не найден.")
        items.append((int(product_row[0]), qty))

    if receive_code is None or not normalize_text(receive_code):
        receive_code_value = int(conn.execute("SELECT COALESCE(MAX(receive_code), 899) + 1 FROM orders").fetchone()[0])
    else:
        receive_code_value = parse_int(receive_code, "Код для получения", 1)

    if order_id is None:
        cur = conn.execute(
            """
            INSERT INTO orders(order_number, status, order_date, delivery_date, receive_code, customer_id, pickup_point_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (order_number_value, status, order_date_value, delivery_date_value, receive_code_value, customer_row[0], pickup_row[0]),
        )
        order_id = int(cur.lastrowid)
    else:
        conn.execute(
            """
            UPDATE orders
            SET order_number = ?, status = ?, order_date = ?, delivery_date = ?, receive_code = ?, customer_id = ?, pickup_point_id = ?
            WHERE id = ?
            """,
            (order_number_value, status, order_date_value, delivery_date_value, receive_code_value, customer_row[0], pickup_row[0], order_id),
        )
        conn.execute("DELETE FROM order_items WHERE order_id = ?", (order_id,))

    for product_id, qty in items:
        conn.execute(
            "INSERT INTO order_items(order_id, product_id, quantity) VALUES (?, ?, ?)",
            (order_id, product_id, qty),
        )
    conn.commit()
    return order_id



def delete_order(conn: sqlite3.Connection, order_id: int) -> None:
    conn.execute("DELETE FROM orders WHERE id = ?", (order_id,))
    conn.commit()



def product_photo_path(photo_path: str) -> Path:
    return _resolve_product_image_path(photo_path)



def safe_open_image(photo_path: str, size: tuple[int, int] = (120, 90)) -> Image.Image:
    path = product_photo_path(photo_path)
    if not path.exists():
        path = PLACEHOLDER_PATH
    with Image.open(path) as img:
        return ImageOps.fit(img.convert("RGB"), size, method=Image.Resampling.LANCZOS)



def copy_import_folder_snapshot(target: Path) -> None:
    if target.exists():
        shutil.rmtree(target)
    shutil.copytree(IMPORT_DIR, target)
